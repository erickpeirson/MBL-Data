
# coding: utf-8

# In[363]:

from openpyxl import load_workbook
from unidecode import unidecode
import matplotlib.pyplot as plt
from collections import Counter
import re
import string
import Levenshtein
import time
from uuid import uuid4


# In[131]:

def strip_punctuation(s):
    exclude = set(string.punctuation)
    s = ''.join(ch for ch in s if ch not in exclude)
    return s


# In[367]:

'{0}'.format(uuid4())


# ## Load course and investigator data

# In[3]:

courselist_wb = load_workbook('/Users/bpeirson/Desktop/MBL Data/MBL COURSE LIST.xlsx')


# In[4]:

investigators_wb = load_workbook('/Users/bpeirson/Desktop/MBL Data/MBL INVESTIGATOR LIST.xlsx')


# In[28]:

coursedata = []
for sheetname in courselist_wb.get_sheet_names():
    sheet = courselist_wb.get_sheet_by_name(sheetname)
    header_row = sheet.rows[0]
    data_rows = sheet.rows[1:]
    for row in data_rows:
        datum = {'year':int(sheetname)}
        for c in xrange(len(header_row)):
            val = row[c].value
            header = header_row[c].value
            if val is not None and header is not None:
                datum[header] = val
        coursedata.append(datum)


# In[340]:

investigatorsdata = []
for sheetname in investigators_wb.get_sheet_names():
    sheet = investigators_wb.get_sheet_by_name(sheetname)
    header_row = sheet.rows[0]
    data_rows = sheet.rows[1:]
    lastrole = None
    lastsubject = None
    for row in data_rows:
        datum = {'year':int(sheetname)}
        for c in xrange(len(header_row)):
            val = row[c].value
            header = header_row[c].value
            if header in ('Independent or Beginner', 'Independent or Beginner?'):
                header = 'Role'
            
            if val is not None and header is not None:
                datum[header] = val
            elif header == 'Role' and val is None:
                val = lastrole
                datum[header] = val
            elif header == 'Subject' and val is None:
                val = lastsubject
                datum[header] = val
                
            if header == 'Role' and val is not None:
                lastrole = unicode(val)
            if header == 'Subject' and val is not None:
                lastsubject = unicode(val)
        investigatorsdata.append(datum)


# In[341]:

investigatorsdata[8092]


# In[248]:

print len(coursedata), len(investigatorsdata), len(coursedata) + len(investigatorsdata)


# ## Normalize course names

# In[153]:

course_names = set([])
f_course_names = Counter()
for datum in coursedata:
    try:
        name = unidecode(datum['Course Name'].strip().lower().replace('&', 'and').replace('1', 'i').replace('2', 'ii'))
        course_names.add(name)
        f_course_names[name] += 1
    except KeyError:
        pass
print len(course_names)


# In[154]:

distances = []
course_names_list = list(course_names)
for i in xrange(len(course_names_list)):
    for j in xrange(i+1, len(course_names_list)):
        a = course_names_list[i]
        b = course_names_list[j]
        d = levenshtein(a,b)
        dnorm = float(d)/mean([float(len(a)), float(len(b))])
        distances.append( (i,j,d, dnorm) )


# In[155]:

for d in distances:
    if d[3] < 0.17:
        a = course_names_list[d[0]]
        b = course_names_list[d[1]]
        dnorm = d[3]

        print d[3], d[2]
        print f_course_names[a], '\t', a
        print f_course_names[b], '\t', b   
        print '-'*40


# In[318]:

# The course_map handles typographical errors in the dataset. There are remarkably
#  few typos. We specify the appropriate spellings manually, below.
course_map = { 
    'optimal microscopy': 'optical microscopy',
    'optimal microscopy and imaging in the biomedical sciences': 'optical microscopy and imaging in the biomedical sciences',
    'nasa planetary biology inernship': 'nasa planetary biology internship',
    'summer program in neuroscience, ethics $ survival': 'summer program in neuroscience, ethics and survival',
    'spines--summer program in neuroscience, ethics and survival': 'summer program in neuroscience, ethics and survival',
    'physiology: modern cell biology using microscopic, biochemical and computational approaches': 'physiology: modern cell biology using microscopic, biochemical, and computational approaches',
    'physiology: cell and molecular biology': 'physiology: cellular and molecular biology',
    'parthogenesis of neuroimmunologic diseases': 'pathogenesis of neuroimmunologic diseases',
}
print len(course_map)


# In[275]:

# If two courses share the same name, then we generally consider them to belong to the 
#  same course group. For example, the 'Ecology' course in 1934 (say) belongs to the
#  same group (or series) as the 'Ecology' course in 1965 (say).
#
# In some cases, however, courses with slightly (or perhaps very) different names
#  may belong to the same group. For example, an Embryology course might have some
#  subtitle, like: "Embryology: Some great new theme for this course". Or they might
#  be numbered, like "Biomedical informatics I" and "Biomedical informatics II".
#
# The coursegroup_map handles the latter cases. Keys are specific course names that
#  occur in the dataset, and values are the group names that should be used.
#  Some of these mappings are given manually, based on inspection of the dataset.
#  Other mappings are generated by looking for colons (':') in course names; the
#  part of the name before the colon is assumed to be the proper group name.

coursegroup_map = { 
    'small computers in biomedical research, i': 'small computers in biomedical research',
    'small computers in biomedical research, ii': 'small computers in biomedical research',
    'medical informatics': 'biomedical informatics',
    'medical informatics i': 'biomedical informatics',
    'medical informatics ii': 'biomedical informatics',
    'biomedical informatics i': 'biomedical informatics',
    'biomedical informatics ii': 'biomedical informatics',
    'advanced workshop on recombinant dna methodology': 'workshop on recombinant dna methodology',
    'basic workshop on recombinant dna methodology': 'workshop on recombinant dna methodology',
}

for cname in list(course_names):   # Here we look for course names with subtitles,
    parts = cname.split(':')       #  characterized by a colon (':') in their names.
    if len(parts) > 1:
        coursegroup_map.update({cname:parts[0]})
print len(coursegroup_map)


# In[297]:

def normalize_coursename(cname):
    cname = unidecode(cname).lower().strip()
    
    if cname in course_map:
        cname = course_map[cname]
    if cname in coursegroup_map:
        group = coursegroup_map[cname]
    else:
        group = cname
    return cname, group


# In[394]:

course_ids = {}
def get_course_uri(cname):
    if cname not in course_ids:
        uri = 'http://history.archives.mbl.edu/concepts/course/{0}'.format(uuid4())  
        course_ids[cname] = uri
    else:
        uri = course_ids[cname]
    return uri


# In[395]:

coursegroup_ids = {}
def get_coursegroup_uri(coursegroup):
    if coursegroup not in coursegroup_ids:
        uri = 'http://history.archives.mbl.edu/concepts/coursegroup/{0}'.format(uuid4())  
        coursegroup_ids[coursegroup] = uri
    else:
        uri = coursegroup_ids[coursegroup]
    return uri


# ## Normalize personal names

# In[383]:

person_ids = {}   # Name -> URI


# In[252]:

personal_names = set([])
f_personal_names = Counter()
personal_affiliations = {}
for datum in coursedata + investigatorsdata:
    try:
        lastname = unidecode(unicode(datum['Last Name']).strip().lower().replace('.',''))
        firstname = unidecode(unicode(datum['First Name']).strip().lower())
        firstname = ' '.join([ n.strip(' ') for n in re.split('\.|\W+', firstname) if n != '' ]).strip().replace('.','')
        affiliations = normalized_institutions(datum['Affiliation'])    # Returns a list.
        name = (lastname, firstname)
        personal_names.add(name)
        f_personal_names[name] += 1
        
        if name not in personal_affiliations:
            personal_affiliations[name] = set([])
        for affiliation in affiliations:  
            personal_affiliations[name].add(affiliation)
    except KeyError:
        pass
    except AttributeError:
        print datum['Last Name']
print len(personal_names)


# In[253]:

personal_names_list = list(personal_names)
N_names = len(personal_names)
by_last = {}
for i in xrange(N_names):
    lastname, firstname = personal_names_list[i]
    if lastname not in by_last:
        by_last[lastname] = set([])
    by_last[lastname].add(firstname)    # The surname of A is identical to the surname of B


# In[378]:

person_map = {}
for last, firsts in by_last.iteritems():    # We assume that surnames are not misspelled.
    N_firsts = len(firsts)                  #  This is not strictly true, but it is not
    if N_firsts > 1:                        #  quite clear how to proceed otherwise.
        lfirsts = list(firsts)              # Consider cases in which two names, I and J,
        for i in xrange(N_firsts):          #  have a common surname.
            iname = lfirsts[i]
            inames = iname.split(' ')
            iinits = [f[0] for f in inames ]
            
            for j in xrange(i+1, N_firsts):     
                jname = lfirsts[j]
                jnames = jname.split(' ')
                jinits = [f[0] for f in jnames ]

                # For each such pair, I and J, we compare the X parts of their forenames,
                #  where X is the minimum number of forename parts for I and J.
                match = True               
                for x in xrange(min( [len(inames), len(jnames)] )):
                    # If the x part if either forename is an initial, we evaluate
                    #  only the first character of the two parts.
                    if len(inames[x]) == 1 or len(jnames[x]) == 1:
                        if iinits[x] != jinits[x]:
                            match = False
                    # Otherwise, the x part of the two forenames must be identical.
                    else:
                        if inames[x] != jnames[x]:
                            match = False
                if match:     
                    # If the forenames of I and J match, as described above, we check
                    # to see whether they share at least one institutional affiliation.
                    shared = personal_affiliations[(last, iname)] & personal_affiliations[(last, jname)]
                    if len(shared) > 0:
                        # If they share at least one institutional affiliation, then
                        #  we believe that I and J both refer to the same person.
                        if len(iname) > len(jname):    # Use the longest name (most complete).
                            key = jname
                            alt = iname
                        else:
                            key = iname
                            alt = jname
                        if (last, alt) in person_map:
                            top = False
                            while not top:
                                try:
                                    alt = person_map[(last,alt)][1]
                                except KeyError:
                                    top = True
                        person_map[(last, key)] = (last, alt)      
                        
    # If the conditions above are not satisfied, then we lack sufficient evidence to
    #  assert that the names I and J refer to the same person.
print len(person_map)


# In[384]:

def normalized_person(last, first):
    """
    Generates a normalized representation of a personal name.
    """
    lastname = unidecode(unicode(last)
                                 .strip()
                                 .lower()
                                 .replace('.',''))
    firstname = unidecode(unicode(first)
                                  .strip()
                                  .lower())
    firstname = ' '.join([ n.strip(' ') for n 
                          in re.split('\.|\W+', firstname) 
                          if n != '' ]).strip().replace('.','')    
            
    name = (lastname, firstname)
    if name in person_map:
        normed_name = person_map[name]
    else:
        normed_name = name
    if normed_name not in person_ids:
        uri = 'http://history.archives.mbl.edu/concepts/person/{0}'.format(uuid4())
        person_ids[normed_name] = uri
    else:
        uri = person_ids[normed_name]
    return normed_name, uri


# In[259]:

i = 0
for datum in investigatorsdata[0:10]:
    if 'Last Name' in datum and 'First Name' in datum:
        print normalized_person(datum['Last Name'], datum['First Name'])
    i += 1
print i


# ## Normalize institution names

# In[249]:

institution_names = set([])
f_institution_names = Counter()
for datum in coursedata + investigatorsdata:
    try:
        affs = unidecode(unicode(datum['Affiliation'])
                             .strip()
                             .lower()
                             .replace('.','')
                             .replace(',','')
                             .replace('&', 'and')
                             .replace('-', ' '))
        for aff in affs.split('/'):
            aff = strip_punctuation(aff.strip()).replace('  ',' ')
            aff = ' '.join([ word for word in aff.split(' ') if word != 'the'])
            institution_names.add(aff)
            f_institution_names[aff] += 1
    except KeyError:
        pass
print len(institution_names)


# In[250]:

institution_distances = []
institution_names_list = list(institution_names)
for i in xrange(len(institution_names_list)):
    for j in xrange(i+1, len(institution_names_list)):
        a = institution_names_list[i]
        b = institution_names_list[j]
        d = Levenshtein.distance(a,b)
        dnorm = float(d)/mean([float(len(a)), float(len(b))])
        institution_distances.append( (i,j,d, dnorm) )


# In[251]:

institutions_lookup = {}
for d in institution_distances:
    if d[3] < 0.1 and d[2] < 4:
        a = institution_names_list[d[0]]
        b = institution_names_list[d[1]]
        f_a = f_institution_names[a]
        f_b = f_institution_names[b]
        
        if f_a > f_b:
            key = a
            alt = b
            f_key = f_a
            f_alt = f_b
        else:
            key = b
            alt = a
            f_key = f_b
            f_alt = f_a
        
        if alt in institutions_lookup:
            f_m = f_institution_names[institutions_lookup[alt]]
            if f_key > f_m:
                institutions_lookup[alt] = key
        else:
            institutions_lookup[alt] = key


# In[388]:

institution_ids = {
    'Marine Biological Laboratory': 'http://history.archives.mbl.edu/concepts/institution/{0}'.format(uuid4()),
}   # Institution -> URI


# In[389]:

def normalized_institutions(inames):
    """
    Generates a normalized representation of an institutional name.
    """
    anames = []
    affs = unidecode(unicode(inames)            # The value of the 'Affiliation' field is
                         .strip()               #  stripped padding whitespace (e.g. spaces),
                         .lower()               #  and converted to lowercase.
                         .replace('&', 'and')   # Ampersands are converted to 'and', and
                         .replace('-', ' '))    #  hyphens are interpreted as spaces.
                                                # The 'Affiliation' field can contain multiple
    for aff in affs.split('/'):                 #  institutions, separated by a slash ('/').
        # All punctuation is removed, and double-spaces are converted to single.     
        aff = strip_punctuation(aff.strip()).replace('  ',' ')  
        aff = ' '.join([ word for word          # One source of variation in names it the
                        in aff.split(' ')       #  inclusion of 'the'. We simply remove
                        if word != 'the'] )     #  'the' from all names.
        
        if aff in institutions_lookup:          # In a previous step, we generated aggregation
            aff = institutions_lookup[aff]      #  rules for some names. If there were multiple
                                                #  similar names, this retrieves the most likely.
        anames.append(aff)
        
        if aff not in institution_ids:
            uri = 'http://history.archives.mbl.edu/concepts/institution/{0}'.format(uuid4())
            institution_ids[aff] = uri            
    return anames 


# In[416]:

location_ids = {}
def get_location_uri(location):
    if location not in location_ids:
        uri = 'http://history.archives.mbl.edu/concepts/location/{0}'.format(uuid4())  
        location_ids[location] = uri
    else:
        uri = location_ids[location]
    return uri


# ## Generate cleaned data

# In[423]:

cleaned_coursedata = []
cleaned_locations = []
cleaned_affiliations = []
cleaned_coursegroups = []
courses_added = set([])
cleaned_investigators = []

for datum in investigatorsdata:
    try:
        datum['Last Name']
        datum['First Name']
    except KeyError:
        continue    
    name, person_uri = normalized_person(datum['Last Name'], datum['First Name'])
    last, first = name

    try:
        location = unidecode(datum['Location']).strip().lower()
        location_uri = get_location_uri(location)
    except KeyError:
        location = ''
    
    try:
        affiliation = datum['Affiliation']
    except KeyError:
        affiliation = None
    
    try:
        role = unidecode(datum['Role'])
    except KeyError:
        role = ''   
    try:
        subject = unidecode(datum['Subject'])
    except:
        subject = ''
        
    try:
        position = unidecode(datum['Position'])
    except KeyError:
        position = ''        
        
    # Person -[hasAffiliation:Position:Year]-> Affiliation
    if affiliation is not None:    # Must have an affiliation.
        for affname in normalized_institutions(affiliation):
            aff_uri = institution_ids[affname]
            cleaned_affiliation = {
                'Person URI': person_uri,
                'Last Name': last.title(),
                'First Name': first.title(),
                'Institution': affname.title(),
                'Institution URI': aff_uri,
                'Position': position,
                'Year': datum['year'],
            }
            cleaned_affiliations.append(cleaned_affiliation)
        
    # Marine Biological Laboratory -[hasInvestigator:Role:Subject:Year]-> Person
    cleaned_investigator = {
        'Role': role.title(),
        'Subject': subject.title(),
        'Year': datum['year'],
        'Person URI': person_uri,
        'First Name': first.title(),
        'Last Name': last.title(),
    }
    cleaned_investigators.append(cleaned_investigator)
        
for datum in coursedata:
    try:
        datum['Last Name']
        datum['First Name']
    except KeyError:
        continue
    name, person_uri = normalized_person(datum['Last Name'], datum['First Name'])
    last, first = name
        
    try:
        cname, coursegroup = normalize_coursename(datum['Course Name'])
        cname = '{0} {1}'.format(cname, datum['year']).title()
        coursegroup = coursegroup.title()
        
        course_uri = get_course_uri(cname)
        coursegroup_uri = get_coursegroup_uri(coursegroup)
    except KeyError:
        cname = None
        coursegroup = None
        
    try:
        position = unidecode(datum['Position at Affiliation'])
    except KeyError:
        position = ''
        
    try:
        location = unidecode(datum['Location']).strip().lower()
        location_uri = get_location_uri(location)        
    except KeyError:
        location = ''
    
    try:
        affiliation = datum['Affiliation']
    except KeyError:
        affiliation = None
    
    try:
        role = unidecode(datum['Role'])
    except KeyError:
        role = ''
        
    # Course -[partOf:Year]-> Group
    if cname not in courses_added:
        cgroup_datum = {
            'Course Group': coursegroup,
            'Course Group URI': coursegroup_uri,
            'Course Name': cname,            
            'Course URI': course_uri,
            'Year': datum['year'],
        }
        cleaned_coursegroups.append(cgroup_datum)
        courses_added.add(cname)
    
    # Person -[hasLocation:Year]-> Location
    cleaned_location = {
        'Person URI': person_uri,
        'Last Name': last.title(),
        'First Name': first.title(),
        'Location': location,    
        'Location URI': location_uri,
        'Year': datum['year'],
    }
    cleaned_locations.append(cleaned_location)
    
    # Person -[attended:Role:Year]-> Course
    if cname is not None:   # Must have a course.
        cleaned_datum = {
            'Course Name': cname,
            'Course URI': course_uri,
            'Role': role,
            'Person URI': person_uri,
            'Last Name': last.title(),
            'First Name': first.title(),
            'Year': datum['year'],
        }
        cleaned_coursedata.append(cleaned_datum)
    else:    # Person -[hasAffiliation:Position:Year]-> "Marine Biological Laboratory"
        aff_uri = institution_ids["Marine Biological Laboratory"]
        cleaned_affiliation = {
            'Person URI': person_uri,
            'Last Name': last.title(),
            'First Name': first.title(),
            'Institution': 'Marine Biological Laboratory',
            'Institution URI': aff_uri,
            'Position': role,
            'Year': datum['year'],
        }
    
    # Person -[hasAffiliation:Position:Year]-> Affiliation
    if affiliation is not None:    # Must have an affiliation.
        for affname in normalized_institutions(affiliation):
            aff_uri = institution_ids[affname]
            cleaned_affiliation = {
                'Person URI': person_uri,
                'Last Name': last.title(),
                'First Name': first.title(),
                'Institution': affname.title(),
                'Institution URI': aff_uri,
                'Position': position,
                'Year': datum['year'],
            }
            cleaned_affiliations.append(cleaned_affiliation)


# In[418]:

print len(coursedata), len(cleaned_coursedata), len(cleaned_affiliations), len(cleaned_coursegroups), len(cleaned_locations), len(cleaned_investigators)


# In[315]:

import csv


# In[398]:

headers = cleaned_coursedata[0].keys()
with open('/Users/bpeirson/Desktop/MBL Data/cleaned_coursedata.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for datum in cleaned_coursedata:
        writer.writerow( [ datum[key] for key in headers ] )


# In[424]:

headers = cleaned_affiliations[0].keys()
with open('/Users/bpeirson/Desktop/MBL Data/cleaned_affiliations.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for datum in cleaned_affiliations:
        writer.writerow( [ datum[key] for key in headers ] )


# In[407]:

headers = cleaned_coursegroups[0].keys()
with open('/Users/bpeirson/Desktop/MBL Data/cleaned_coursegroups.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for datum in cleaned_coursegroups:
        writer.writerow( [ datum[key] for key in headers ] )


# In[419]:

headers = cleaned_locations[0].keys()
with open('/Users/bpeirson/Desktop/MBL Data/cleaned_locations.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for datum in cleaned_locations:
        writer.writerow( [ datum[key] for key in headers ] )


# In[409]:

headers = cleaned_investigators[0].keys()
with open('/Users/bpeirson/Desktop/MBL Data/cleaned_investigators.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for datum in cleaned_investigators:
        writer.writerow( [ datum[key] for key in headers ] )


# In[412]:

with open('/Users/bpeirson/Desktop/MBL Data/person.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerow(['Last Name', 'First Name', 'Person URI'])
    for name, uri in person_ids.iteritems():
        last,first = name
        writer.writerow([last.title(), first.title(), uri])


# In[414]:

with open('/Users/bpeirson/Desktop/MBL Data/institution.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerow(['Institution', 'Institution URI'])
    for institution, uri in institution_ids.iteritems():
        writer.writerow([institution.title(), uri])


# In[420]:

with open('/Users/bpeirson/Desktop/MBL Data/location.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerow(['Location', 'Location URI'])
    for location, uri in location_ids.iteritems():
        writer.writerow([location, uri])


# In[421]:

with open('/Users/bpeirson/Desktop/MBL Data/course.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerow(['Course Name', 'Course URI'])
    for cname, uri in course_ids.iteritems():
        writer.writerow([cname, uri])


# In[422]:

with open('/Users/bpeirson/Desktop/MBL Data/coursegroup.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerow(['Course Group', 'Course Group URI'])
    for cname, uri in coursegroup_ids.iteritems():
        writer.writerow([cname, uri])


# In[ ]:



